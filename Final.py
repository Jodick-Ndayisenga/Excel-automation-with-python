
from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
#FONT1 = Font(name='Times New Roman', size=13, bold=True)

wb = load_workbook('Registered Students.xlsx')
CLASSES = wb.sheetnames
''''for bad in wb.sheetnames:
    bd = wb[bad]
    bd.delete_rows(1,100)
    bd.append(['NUMBER', 'STUDENT NAME', 'STUDENT ID'])
    for col in range(1,4):
            bd[get_column_letter(col)+'1'].font = Font(bold=True, italic=True, size=9)
wb.save('Registered Students.xlsx')
print('Deal is over bro!!')'''''''''


courses = {
    'SUS1010': 'Strategy for University Success',
    'IST1020': 'Introduction to Information Systems',
    'MTH1109': 'College Algebra',
    'MTH1110': 'Calculus',
    'IST1025': 'Introduction to Programming',
    'ENG1106': 'Composition I',
    'MTH2215': 'Discrete Mathematics',
    'APT1030': 'Fundamentals of Programming Languages',
    'APT1040': 'Introduction to Web Design & Applications',
    'APT1050': 'Database Systems',
    'APT2010': 'Systems Analysis & Design',
    'APT2020': 'Computer Organization',
    'ENG2206': 'Composition II',
    'GRM2000': 'Introduction to Research Methods',
    'APT2022': 'Introduction to Assembly Programming',
    'APT2030': 'Digital Electronics',
    'APT2040': 'Operating Systems',
    'IST2045': 'Introduction to Computer Networks',
    'APT2050': 'Computer Networks & Telecommunications',
    'APT2055': 'Hardware and Software Practicum',
    'APT2060': 'Data Structures & Algorithms',
    'APT2080': 'Introduction to Software Engineering',
    'CMS3700': 'Community Service',
    'APT2090': 'Computer Graphics',
    'IST3015': 'Business Data Analytics',
    'IST3050': 'Introduction to Security Systems',
    'APT3010': 'Introduction to Artificial Intelligence',
    'APT3025': 'Applied Machine Learning',
    'APT3040': 'Object Oriented Analysis, Design & Programming',
    'APT3050': 'Introduction to Project Management',
    'APT3060': 'Mobile Programming',
    'APT3080': 'Management Information Systems',
    'IST4035': 'Advanced Web Design and Applications',
    'APT3090': 'Cryptography & Network Security',
    'APT3095': 'Cloud Computing and Visualization',
    'IST4078': 'IT Innovation and Entrepreneurship',
    'SEN4800': 'Integrated Senior Seminar',
    'APT4900': 'Applied Computer Technology Project',
    'APT4910': 'Applied Computer Technology Internship',
    'DST4010': 'Distributed Systems',
    'NET4020': 'Netwok Management & Security',
    'NET4030': 'Internet Architecture & Protocols',
    'APP4035': 'Web Applications Development Frameworks',
    'NET4050': 'Sensor Networks & Internet of Things',
    'FIC4010': 'Information System Security',
    'FIC4020': 'Forensic Accounting & Fraud Investigation Security',
    'FIC4030': 'Information Systems Audit',
    'FIC4040': 'Information Technology & Cybercrime',
    'FIC4050': 'Computer Forensics & Investigation',
    'SFE4010': 'Human Computer Interaction',
    'SFE4020': 'Software Design & Architecture',
    'SFE4030': 'Software Testing & Quality Assurance',
    'APP4035': 'Web Application Development Frameworks',
    'APP4080': 'Collaborative (Team Work In) Software Development'}


def Register(x):
    sheet = wb[x]
    sheet['A2'] = 0
    RANGE = _Class['A']
    for cell in RANGE:
        if cell.value != None:
            number = cell.value
    number += 1
    sheet.append([number, Name, ID])
    wb.save('Registered Students.xlsx')


def Checking(x='ENG0999'):
    ''''workSheet=wb[classToRegister]
    studentIdentity=workSheet['C']
    for cell1 in studentIdentity:
        if ID==cell1.value:
            print(f'{Name}, you have already registered for this class!')
            break       
        else:'''
    try:
        score = input(
            f'What is the average score did you get in {x} or placement test?')
        score = int(score)

        if score <= 69:
            print(
                f'We are sorry,{Name}. You cannot register for this class with that score')
            print(
                'Please, try to register for introduction to information system technology remedial')
        else:
            Register(classToRegister)
            print('')
    except:
        print('The score has to be in numbers!')
        print('')


while True:
    Name = str(input('What is your full name? '))
    ID = input('What is your student ID? ')

    classToRegister = str(
        ((input('What is the class you want to register for? ')).strip()).upper())

    if classToRegister in CLASSES:
        _Class = wb[classToRegister]

        if classToRegister == 'IST1025':
            Checking('IST1020')

        elif classToRegister == 'SUS1010':
            Checking()

        elif classToRegister == 'ENG1106':
            Checking('ENG0999')

        elif classToRegister == 'MTH1109':
            Checking('MTH1105')

        elif classToRegister == 'APT1030':
            Checking('IST1025')

        elif classToRegister == 'IST1020':
            Checking('IST0999')

        elif classToRegister == 'MTH1110':
            Checking('MTH1109')

        elif classToRegister == 'MTH2215':
            Checking('MTH1109')

        elif classToRegister == 'APT1040':
            Checking('IST1020')

        elif classToRegister == 'APT1050':
            Checking('IST1025')

        elif classToRegister == 'APT2010':
            Checking('IST1020')

        elif classToRegister == 'APT2020':
            Checking('IST1025')

        elif classToRegister == 'ENG2206':
            Checking('ENG1106')

        elif classToRegister == 'GRM2000':
            Checking('SUS1010')

        elif classToRegister == 'APT2022':
            Checking('APT1030')

        elif classToRegister == 'APT2030':
            Checking('APT2020')

        elif classToRegister == 'APT2040':
            Checking('APT2020')

        elif classToRegister == 'IST2045':
            a = 'APT1030 or ENG2206'
            Checking(a)

        elif classToRegister == 'APT2050':
            Checking('APT2020')

        elif classToRegister == 'APT2O55':
            Checking('IST1045')

        elif classToRegister == 'APT2060':
            b = 'MTH2215 or APT1030'
            Checking(b)

        elif classToRegister == 'APT2080':
            Checking('IST1020')

        elif classToRegister == 'CMS3700':
            Checking()

        elif classToRegister == 'APT2090':
            Checking('APT2060')

        elif classToRegister == 'APT3015':
            Checking('MTH15')

        elif classToRegister == 'APT3050':
            Checking('APT2050')

        elif classToRegister == 'APT3010':
            Checking('APT2090')

        elif classToRegister == 'APT3025':
            Checking('APT3010')

        elif classToRegister == 'APT3040':
            Checking('APT1030')

        elif classToRegister == 'APT2090':
            Checking('APT2060')

        elif classToRegister == 'IST3015':
            Checking('MTH2215')

        elif classToRegister == 'IST3050':
            Checking('APT2015')

        elif classToRegister == 'APT3010':
            Checking('APT2090')

        elif classToRegister == 'APT3025':
            Checking('APT3010')

        elif classToRegister == 'APT3040':
            Checking('APT1030')

        elif classToRegister == 'APT3050':
            Checking('APT2010')

        elif classToRegister == 'APT3060':
            Checking('APT3040')

        elif classToRegister == 'APT3065':
            f = 'APT3060 or CMS3700'
            Checking(f)

        elif classToRegister == 'APT3080':
            Checking('IST1025')

        elif classToRegister == 'IST4035':
            Checking('APT1040')

        elif classToRegister == 'APT3090':
            Checking('APT2050')

        elif classToRegister == 'APT3095':
            Checking('APT2050')

        elif classToRegister == 'IST4078':
            Checking('APT3060')

        elif classToRegister == 'SEN4800':
            Checking('IST4078')

        elif classToRegister == 'APT4900':
            Checking('APT3065')

        elif classToRegister == 'APT4910':
            Checking('APT3090')
    else:
        print(
            f'Sorry,{Name} the class you are trying to register for is not available this semester')
        print('')
